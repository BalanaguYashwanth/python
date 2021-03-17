import pandas as pd
import openpyxl as xl;
                
data=pd.read_excel(r'C:\Users\KZ623JK\Downloads\output.xlsx')
data.to_excel('newfile.xlsx',index=False)

wb=xl.load_workbook(r'C:\Users\KZ623JK\newfile.xlsx')
sheet=wb.worksheets[0]
ws=wb.active

mr=sheet.max_row
mc=sheet.max_column
#print(mr,mc) #apply these mr,mc in condition logic

a=input('enter start date')
b=input('enter end date')

a1=int(a.split('/')[1])
b1=int(b.split('/')[1])

c=b1-a1

for x in range(c+1):
    print(x,c,a1)
    ws.cell(row = 1, column = x+1).value = str(a1+x)+'-march-'+'2021'
    #ws.cell(0,x+1).value = a1+x

wb.save(r'C:\Users\KZ623JK\newfile.xlsx')
