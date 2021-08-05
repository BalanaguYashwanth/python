[20:45] Jayshree Tomar
    
import openpyxl
import re


from openpyxl.xml.constants import MAX_COLUMN, MIN_COLUMN
wb=openpyxl.load_workbook("F:\\asads.xlsx")
sheets=wb.sheetnames
sh1=wb['Carrier_CommercialAuto_Forms_Mu']
#print(sh1.cell(1,8).value)
row=sh1.max_row
column=sh1.max_column
sh1.cell(row=1,column=12,value='Issue New')
sh1.cell(row=1,column=13,value='Bind')
sh1.cell(row=1,column=14,value='IssueIssuePolicy')
sh1.cell(row=1,column=15,value='IssueRewrite')
sh1.cell(row=1,column=16,value='IssueRenewal')
sh1.cell(row=1,column=17,value='IssueReissue')
sh1.cell(row=1,column=18,value='IssueEndorse')
sh1.cell(row=1,column=19,value='IssueCancel')
sh1.cell(row=1,column=20,value='ssueReinstate')
sh1.cell(row=1,column=21,value='IssueRescind')
sh1.cell(row=1,column=22,value='Static/Variable')
sh1.cell(row=1,column=23,value='# of Variable Fields')



row = int(5)
j = int(8)
for i in range(2,row):
    #for j in range(8,column-2):
        mystr = (sh1.cell(i,j).value)
       
        if mystr!=None: 
            print(mystr)
            print(i)
            sh1.cell(row=i,column=22,value='Variable')
            
        else:
            sh1.cell(row=i,column=22,value='static')
wb.save("port.xlsx")              











 
import openpyxl
import re


from openpyxl.xml.constants import MAX_COLUMN, MIN_COLUMN
wb=openpyxl.load_workbook("F:\\asads.xlsx")
sheets=wb.sheetnames
sh1=wb['Carrier_CommercialAuto_Forms_Mu']
#print(sh1.cell(1,8).value)
row=sh1.max_row
column=sh1.max_column
sh1.cell(row=1,column=12,value='Issue New')
sh1.cell(row=1,column=13,value='Bind')
sh1.cell(row=1,column=14,value='IssueIssuePolicy')
sh1.cell(row=1,column=15,value='IssueRewrite')
sh1.cell(row=1,column=16,value='IssueRenewal')
sh1.cell(row=1,column=17,value='IssueReissue')
sh1.cell(row=1,column=18,value='IssueEndorse')
sh1.cell(row=1,column=19,value='IssueCancel')
sh1.cell(row=1,column=20,value='ssueReinstate')
sh1.cell(row=1,column=21,value='IssueRescind')
sh1.cell(row=1,column=22,value='Static/Variable')
sh1.cell(row=1,column=23,value='# of Variable Fields')



row = int(5)
j = int(8)
for i in range(2,row):
    #for j in range(8,column-2):
        mystr = (sh1.cell(i,j).value)
       
        if mystr!=None: 
            print(mystr)
            print(i)
            sh1.cell(row=i,column=22,value='Variable')
            
        else:
            sh1.cell(row=i,column=22,value='static')
wb.save("port.xlsx")              



















import openpyxl
import re


from openpyxl.xml.constants import MAX_COLUMN, MIN_COLUMN
wb=openpyxl.load_workbook("F:\\asads.xlsx")
sheets=wb.sheetnames
sh1=wb['Carrier_CommercialAuto_Forms_Mu']
#print(sh1.cell(1,8).value)
row=sh1.max_row
column=sh1.max_column
sh1.cell(row=1,column=12,value='Issue New')
sh1.cell(row=1,column=13,value='Bind')
sh1.cell(row=1,column=14,value='IssueIssuePolicy')
sh1.cell(row=1,column=15,value='IssueRewrite')
sh1.cell(row=1,column=16,value='IssueRenewal')
sh1.cell(row=1,column=17,value='IssueReissue')
sh1.cell(row=1,column=18,value='IssueEndorse')
sh1.cell(row=1,column=19,value='IssueCancel')
sh1.cell(row=1,column=20,value='ssueReinstate')
sh1.cell(row=1,column=21,value='IssueRescind')
sh1.cell(row=1,column=22,value='Static/Variable')
sh1.cell(row=1,column=23,value='# of Variable Fields')



row = int(5)
j = int(8)
for i in range(2,row):
    #for j in range(8,column-2):
        mystr = (sh1.cell(i,j).value)
       
        if mystr!=None: 
            print(mystr)
            print(i)
            sh1.cell(row=i,column=22,value='Variable')
            
        else:
            sh1.cell(row=i,column=22,value='static')
wb.save("port.xlsx")              




















