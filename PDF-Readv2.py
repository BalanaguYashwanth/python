!pip install PyPDF2

import os
import PyPDF2 as pypdf
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\KZ623JK\OneDrive - EY\Documents\Sample-Data.xlsx")
sheet = wb.active
directory=input('enter the file directory')
row=1

for file in os.scandir(directory):
    pdfobject=open(file.path,'rb')
    pdf=pypdf.PdfFileReader(pdfobject)
    row=row+1
    arr=[]
    arrexp=[]
    arrtags=[]
    arrlocation=[]
    for num in range(pdf.numPages):
        flag=0
        for x in pdf.getPage(num).extractText().splitlines():    
            
            if len(arr)<10:
                arr.append(x)
            sheet.cell(row=row,column=1).value="".join(arr)
            if 'experience' in x:
                arrexp.append(x)
            sheet.cell(row=row,column=7).value="".join(arrexp)
            if 'integration' in x.lower() or 'billing' in x.lower() or 'policy' in x.lower() or 'insights' in x.lower() or 'duckcreek' in x.lower():
                print(x)
            sheet.cell(row=row,column=4).value="".join(arrexp)
            if 'location' in x.lower():
                print(x)
            sheet.cell(row=row,column=5).value="".join(arrexp)
    wb.save(r"C:\Users\KZ623JK\OneDrive - EY\Documents\Sample-Data.xlsx")

print('completed')
